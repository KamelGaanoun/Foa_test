import openpyxl
from openpyxl.utils import get_column_letter 
from openpyxl.drawing.image import Image
from PIL import Image as PILImage
import io
import os
import streamlit as st
from collections import defaultdict
import re
import zipfile

# Function to resize photos to a fixed size
# def resize_photo(image_data, target_width_cm, target_height_cm):
#     # Convert centimeters to pixels (1 cm ≈ 37.795 pixels)
#     target_width_px = int(target_width_cm * 37.795)
#     target_height_px = int(target_height_cm * 37.795)

#     # Open the image using PIL
#     pil_img = PILImage.open(io.BytesIO(image_data))
#     # Resize the image to the target size
#     resized_img = pil_img.resize((target_width_px, target_height_px), PILImage.ANTIALIAS)
#     return resized_img


# Function to enforce consistent row increments
# def enforce_consistent_row_increments(sheet):
#     expected_row = 1  # Start with the first row
#     for img in sheet._images:
#         anchor_cell = img.anchor._from
#         current_row = anchor_cell.row

#         # If the current row is not the expected row, adjust it
#         if current_row != expected_row:
#             anchor_cell.row = expected_row  # Update the row position

#         # Increment the expected row by 2
#         expected_row += 2

# Function to extract and print positions
# def extract_and_print_positions(sheet):
#     # Extract old image positions
#     print("### Old Image Positions")
#     old_image_positions = []
#     for idx, img in enumerate(sheet._images):
#         anchor_cell = img.anchor._from
#         row = anchor_cell.row
#         col = anchor_cell.col
#         old_image_positions.append((row, col, img))
#         print(f"Image {idx + 1}: Row {row}, Column {col}")

#     # Sort images by row and column (convert column to one-based indexing)
#     old_image_positions.sort(key=lambda x: (x[0], x[1] + 1))  # Add 1 to column index

#     # Extract text positions (skip the first row)
#     print("### Text Positions")
#     text_positions = []
#     for row in sheet.iter_rows(min_row=2):  # Start from the second row
#         for cell in row:
#             if cell.value and isinstance(cell.value, str):  # Check if the cell contains text
#                 text_positions.append((cell.row, cell.column, cell.value))
#                 print(f"Text: '{cell.value}' at Row {cell.row}, Column {cell.column}")

#     return old_image_positions, text_positions

# Function to resize photos within the Excel file and print positions
# def resize_photos_in_excel(file_path, target_width_cm, target_height_cm):
#     # Load the workbook and select the active sheet
#     workbook = openpyxl.load_workbook(file_path)
#     sheet = workbook.active

#     # Enforce consistent row increments for images
#     enforce_consistent_row_increments(sheet)

#     # Extract and print old image positions and text positions
#     old_image_positions, text_positions = extract_and_print_positions(sheet)

#     # Create a list to store resized images and their anchors
#     resized_images = []

#     # Iterate through the sorted images
#     for row, col, img in old_image_positions:
#         # Access the image data by calling the _data() method
#         image_data = img._data()  # Call _data() to get the image data

#         # Resize the photo to fit the cell
#         resized_photo = resize_photo(image_data, target_width_cm, target_height_cm)

#         # Convert the resized photo to bytes
#         img_bytes = io.BytesIO()
#         resized_photo.save(img_bytes, format="PNG")
#         img_bytes.seek(0)

#         # Store the resized image and its anchor
#         resized_images.append((img.anchor, img_bytes))

#     # Delete all old images
#     for idx in reversed(range(len(sheet._images))):
#         del sheet._images[idx]

#     # Insert the resized images
#     for anchor, img_bytes in resized_images:
#         resized_img = Image(img_bytes)
#         sheet.add_image(resized_img, anchor=anchor)

#     # Extract and print new image positions
#     print("### New Image Positions")
#     new_image_positions = []
#     for idx, img in enumerate(sheet._images):
#         anchor_cell = img.anchor._from
#         row = anchor_cell.row
#         col = anchor_cell.col
#         new_image_positions.append((row, col))
#         print(f"Image {idx + 1}: Row {row}, Column {col}")
#     # Save the updated workbook
#     output_file = file_path.replace(".xlsx", "_resized.xlsx")
#     workbook.save(output_file)
#     workbook.close()



############################################################
def get_adjusted_row(img):
    """Estimate the closest row by rounding image anchor position."""
    return round(img.anchor._from.row + img.anchor._from.rowOff / (18.75 * 9525))


def extract_text_positions(sheet):
    """
    Extracts text from columns 1 and 2 and generates missing text for columns 3 and 4.
    Returns a dictionary mapping (row, col) to generated image names.
    """
    text_positions = {}

    for row in range(3, sheet.max_row + 1):  # Ignore header rows
        for col in [1, 2]:  # Only process columns 1 and 2
            cell_value = sheet.cell(row=row, column=col).value

            if cell_value and isinstance(cell_value, str):  # Ensure valid text
                base_name, _ = os.path.splitext(cell_value)[0].split('_')[0],os.path.splitext(cell_value)[1] # Extract base name
                
                # Store original text with expected image name
                text_positions[(row, col)] = f"{base_name}_1.jpg" if col == 1 else f"{base_name}_2.jpg"

                # Generate new text values for columns 3 and 4
                text_positions[(row, 3)] = f"{base_name}_3.jpg"
                text_positions[(row, 4)] = f"{base_name}_4.jpg"

    return text_positions


def assign_columns_by_visual_order(sheet):
    """
    Assign column indices to images based on their visual order in the sheet.
    """
    # Create a dictionary to store images by row
    images_by_row = defaultdict(list)

    # Group images by row
    for img in sheet._images:
        img_row = get_adjusted_row(img)
        images_by_row[img_row].append(img)

    # Assign column indices based on visual order
    image_positions = []
    for row, images in sorted(images_by_row.items()):
        # Sort images in the row by their horizontal position
        images.sort(key=lambda img: img.anchor._from.col)
        for col, img in enumerate(images, start=1):  # Start column index at 1
            image_positions.append((row, col, img))

    return image_positions

# #V1
# def extract_images_with_correct_names(excel_file, output_folder):
#     """
#     Extract images from an Excel sheet and name them based on the text at (image_row + 2, image_col + 1).
#     Adjust duplicate names (_1 → _3, _2 → _4) where applicable and fill missing text names.
#     """
#     wb = openpyxl.load_workbook(excel_file)
#     sheet = wb['Photos']  # Use the active sheet or specify another one

#     os.makedirs(output_folder, exist_ok=True)  # Ensure the output folder exists

#     # Step 1: Extract all images and their positions
#     image_positions = []
#     for img in sheet._images:
#         img_row = get_adjusted_row(img)  # Adjusted row position
#         img_col = img.anchor._from.col
#         #img_col=get_adjusted_col(img,sheet)
#         image_positions.append((img_row, img_col))
#     # Step 2: Extract text values, ignoring rows < 3
#     text_positions = extract_text_positions(sheet)
    
    
#     # Step 5: Assign the correct name to each image
#     for img, (img_row, img_col) in zip(sheet._images, image_positions):
#         text_position = (img_row + 2, img_col + 1)  # Get the expected text location
#         image_name = text_positions.get(text_position, None)

#         if image_name:
#             img_bytes = img.ref.getvalue()  # Extract binary image data
#             pil_image = PILImage.open(io.BytesIO(img_bytes))  # Convert to PIL Image


#             # Ensure the filename does not have an extra extension
#             base_name, _ = os.path.splitext(image_name)  # Remove existing extension
#             img_filename = f"{base_name}.png"  # Use only the base name + .png

            
#             #img_filename = f"{image_name}.png"
#             img_filepath = os.path.join(output_folder, img_filename)
#             pil_image.save(img_filepath)  # Save image with assigned name

#             print(f"✅ Image saved as {img_filename} (from row {img_row})")
#         else:
#             print(f"⚠ No text found for image at (Row {img_row}, Col {img_col})")

#     #wb.save(excel_file)  # Save the updated Excel file with added texts
#     wb.close()
#     print(f"✅ All images extracted and saved in: {output_folder}")


# #V2
# def extract_images_with_correct_names(excel_file, output_folder):
#     """
#     Extract images from an Excel sheet and name them based on the text at (image_row + 2, image_col + 1).
#     Adjust duplicate names (_1 → _3, _2 → _4) where applicable and fill missing text names.
#     """
#     wb = openpyxl.load_workbook(excel_file)
#     sheet = wb['Photos']  # Use the active sheet or specify another one

#     os.makedirs(output_folder, exist_ok=True)  # Ensure the output folder exists

#     # Step 1: Extract all images and their positions
#     image_positions = []
#     for img in sheet._images:
#         img_row = get_adjusted_row(img)  # Adjusted row position
#         img_col = img.anchor._from.col
#         image_positions.append((img_row, img_col))

#     # Step 2: Extract text values, ignoring rows < 3
#     text_positions = extract_text_positions(sheet)

#     # Step 3: Dictionary to track assigned names
#     assigned_names = {}

#     # Step 4: Assign the correct name to each image
#     for img, (img_row, img_col) in zip(sheet._images, image_positions):
#         text_position = (img_row + 2, img_col + 1)  # Expected text location
#         image_name = text_positions.get(text_position, None)

#         if image_name:
#             # Remove any existing file extension
#             base_name, _ = os.path.splitext(image_name)

#             # Handle naming conflicts (_1 → _3, _2 → _4)
#             if base_name in assigned_names:
#                 if assigned_names[base_name].endswith("_1.png"):
                    
#                     img_filename = f"{base_name.split('_')[0]}_3.png"
#             else:
#                 # Default naming (_1 or _2 based on column)
#                 #img_filename = f"{base_name}_1.png" if img_col in [1, 3] else f"{base_name}_2.png"
#                 img_filename = f"{base_name}.png"
#             # Save assigned name to avoid duplicates
#             assigned_names[base_name] = img_filename

#             # Save the image
#             img_filepath = os.path.join(output_folder, img_filename)
#             img_bytes = img.ref.getvalue()  # Extract binary image data
#             pil_image = PILImage.open(io.BytesIO(img_bytes))
#             pil_image.save(img_filepath)

#             print(f"✅ Image saved as {img_filename} (from row {img_row})")
#         else:
#             print(f"⚠ No text found for image at (Row {img_row}, Col {img_col})")


#V3
def extract_base_name_and_suffix(filename):
    """
    Extract the base name and suffix from a filename.
    Example: "1215545_2.jpg" → ("1215545", "_2")
    """
    # Remove the file extension (e.g., ".jpg", ".png")
    filename_without_ext = os.path.splitext(filename)[0]

    # Extract base name and suffix
    match = re.match(r"(.+?)(_(\d+))?$", filename_without_ext)
    if match:
        base_name = match.group(1)
        suffix = f"_{match.group(3)}" if match.group(3) else ""
        return base_name, suffix
    return filename_without_ext, ""

def get_next_available_suffix(output_folder, base_name):
    """
    Find the next available suffix for a base name in the output folder.
    Example: If "1215545_1.png" and "1215545_2.png" exist, return "_3".
    """
    existing_files = os.listdir(output_folder)
    existing_suffixes = []

    for file in existing_files:
        if file.startswith(base_name):
            file_base, file_suffix = extract_base_name_and_suffix(file)
            if file_base == base_name and file_suffix:
                existing_suffixes.append(int(file_suffix[1:]))  # Remove "_" and convert to int

    if not existing_suffixes:
        return "_1"  # If no files exist, start with "_1"

    # Find the smallest missing suffix
    existing_suffixes.sort()
    for i, suffix in enumerate(existing_suffixes, start=1):
        if i != suffix:
            return f"_{i}"  # Return the first missing suffix

    return f"_{len(existing_suffixes) + 1}"  # If no gaps, return the next suffix

def extract_images_with_correct_names(excel_file, output_folder):
    """
    Extract images from an Excel sheet and name them based on the text at (image_row + 2, image_col + 1).
    Handle duplicate names by assigning the next available suffix.
    """
    wb = openpyxl.load_workbook(excel_file)
    sheet = wb["Photos"]  # Use the active sheet or specify another one

    os.makedirs(output_folder, exist_ok=True)  # Ensure the output folder exists

    # Step 1: Extract all images and their positions
    image_positions = []
    for img in sheet._images:
        img_row = get_adjusted_row(img)  # Adjusted row position
        img_col = img.anchor._from.col  # Adjusted column position
        image_positions.append((img_row, img_col))

    # Step 2: Extract text values, ignoring rows < 3
    text_positions = extract_text_positions(sheet)

    # Step 3: Assign the correct name to each image
    for img, (img_row, img_col) in zip(sheet._images, image_positions):
        text_position = (img_row + 2, img_col + 1)  # Get the expected text location
        image_name = text_positions.get(text_position, None)

        if image_name:
            img_bytes = img.ref.getvalue()  # Extract binary image data
            pil_image = PILImage.open(io.BytesIO(img_bytes))  # Convert to PIL Image

            # Extract base name and suffix
            base_name, suffix = extract_base_name_and_suffix(image_name)

            # Check if the file already exists
            if os.path.exists(os.path.join(output_folder, f"{base_name}{suffix}.png")):
                # Find the next available suffix
                suffix = get_next_available_suffix(output_folder, base_name)

            # Save the image with the new suffix
            img_filename = f"{base_name}{suffix}.png"
            img_filepath = os.path.join(output_folder, img_filename)
            pil_image.save(img_filepath)  # Save image with assigned name

            print(f"✅ Image saved as {img_filename} (from row {img_row})")
        else:
            print(f"⚠ No text found for image at (Row {img_row}, Col {img_col})")

    wb.close()
    print(f"✅ All images extracted and saved in: {output_folder}")

    # Step 4: Create a ZIP archive of the extracted images
    zip_filename = os.path.join(output_folder, "extracted_images.zip")
    with zipfile.ZipFile(zip_filename, "w") as zipf:
        for root, _, files in os.walk(output_folder):
            for file in files:
                if file.endswith(".png"):
                    file_path = os.path.join(root, file)
                    zipf.write(file_path, os.path.basename(file_path))

    return zip_filename














