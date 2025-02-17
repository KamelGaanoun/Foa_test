
import streamlit as st
import  streamlit_toggle as tog
from streamlit_option_menu import option_menu
import openpyxl as px
from C6_extract import extract_images_with_correct_names
from C6_insert import extract_photo_names_from_excel, filter_photo_names, find_matching_photos,process_excel
from htmlTemplates import css,upload_style
from io import BytesIO
from openpyxl.drawing.image import Image
from PIL import Image as PILImage
import zipfile

import json
import shutil
import os
from reading import load_excel_file
from uploading import upload_files
from regEx import get_planche_number,process_text
from insert import foa_feeder
from C6allignment import extract_images,extract_texts,allign_images,flatten_images,flatten_texts, save_images_to_zip
import tempfile



def main():


    st.set_page_config(page_title="FOA Builder", page_icon=":satellite_antenna:")
    
    st.sidebar.markdown("""
    <style>
        .sidebar .sidebar-content {
            background-color: #F8F9FA;  /* Light grey background */
            padding: 10px;
        }
        .sidebar .sidebar-content a {
            font-size: 18px;
            color: #333;
            text-decoration: none;
            display: flex;
            align-items: center;
            padding: 10px 15px;
            border-radius: 5px;
        }
        .sidebar .sidebar-content a:hover {
            background-color: #EAEAEA;
        }
        .icon {
            width: 20px;
            margin-right: 10px;
        }
    </style>
    """, unsafe_allow_html=True)
    

    st.markdown("""
    <style>
    [data-testid=stSidebar] {
        background-color:  #ffa500;
        border-radius: 0px !important; /* Removes rounded corners */
        padding: 0px !important; /* Removes internal spacing */
        margin: 0px !important; /* Removes external spacing */
    }
    </style>
    """, unsafe_allow_html=True)

    with st.sidebar:        
            app = option_menu(
                menu_title='Opérations',
                options=["FOA", "Annexe C6"],
                icons=['',''],
                menu_icon='gear',
                default_index=0,
                styles={ "menu-title":{"color": "black", "font-size": "18px", "font-weight": "bold"},  	
                    "container": {"margin": "0!important","border-radius": "0!important","padding": "0!important","background-color":'orange'},
        "icon": {"color": "black", "font-size": "13px"}, 
        "nav-link": {"color":"black","font-size": "13px", "text-align": "left", "margin":"0px", "--hover-color": "navajowhite"},
        "nav-link-selected": {"background-color": "#ff8c00"},
        "nav-link-hover": { "text-decoration": "underline" }}
                
                )

    
    if app == "FOA":
        
        
        # Inject the custom CSS to override Streamlit file uploader
        st.markdown(upload_style,unsafe_allow_html=True,)
        st.write(css, unsafe_allow_html=True)

        st.markdown(f'<h1 style="color:#ffa500;font-size:44px;">{"FOA"}</h1>', unsafe_allow_html=True)
        #st.header("FOA")


        #st.header("Créez votre FOA en un clic :satellite_antenna:")


        
        #Uploading the file
        uploaded_files=upload_files(type="xlsx",multiple=False)


        #Toggle to save images or not
        image_save=st.toggle("Voulez-vous sauvegarder les photos ?",value=False)


        if uploaded_files:
            #Gettin workbook from the session, otherwise st keeps refreshing and reload 
            #the wb each time it refreshes, and decreases user exp quality
            if "workbook" not in st.session_state:
                with st.spinner("Lecture et validation du fichier..."):
                    wb = load_excel_file(uploaded_files,read_only=True)  # Load the Excel file
                    st.session_state.workbook = wb  # Store the workbook in session state
                    st.session_state.sheet_names = wb.sheetnames  # Store sheet names

            # Access stored workbook and sheet names
            wb = st.session_state.workbook
            sheet_names = st.session_state.sheet_names

            # Add an empty default choice (So the user is encouraged to correctly choose
            #the correct sheet name, otherwise maybe he will inadvertly let the first option
            #on the menu, which is not surely the good one)
            sheet_options = [""] + sheet_names  # Prepend an empty choice

            #Get nmero commande
            command_num=st.text_input("Merci de renseigner le numéro de commande")        
            
            if command_num:
                # Step 2: Prompt for the sheet name and provide a button to process
                sheet_name = st.selectbox("Veuillez sélectionner la feuille contenant les photos puis cliquez sur Traiter", sheet_options)
                process_button = st.button("Traiter")

                wb.close()  # Close the workbook to free up memory

                #What happens when "Traiter" is clicked
                if process_button:
                    # Check if the user has selected a valid sheet name
                    if sheet_name == "":
                        st.warning("Veuillez sélectionner une feuille avant de continuer.")
                    else:
                        with st.spinner("Traitement en cours..."):
                            try:
                                wb = load_excel_file(uploaded_files,read_only=False)
                                #get the selected sheet
                                sheet = wb[sheet_name]

                                # Store PIL images in memory
                                pil_images_in_memory = []

                                # Create a directory with the same name as the uploaded file (without the extension)
                                base_filename = os.path.splitext(uploaded_files.name)[0]  # Remove .xlsx extension
                                save_directory = os.path.join("Images", base_filename)
                                outputs_directory = os.path.join("Outputs", base_filename)
                                for directory in [save_directory, outputs_directory]:
                                    os.makedirs(directory, exist_ok=True)

                                    
                                
                                # Extract images along with their positions (row and column)
                                # images_with_positions = []
                                # for image in sheet._images:
                                #     # Extract row and column from the image's anchor
                                #     row = image.anchor._from.row
                                #     col = image.anchor._from.col
                                #     images_with_positions.append((row, col, image))

                                # # Sort images by row (top-to-bottom) and column (left-to-right)
                                # images_with_positions.sort(key=lambda x: (x[0], x[1]))

                                images_with_positions = sorted(
                                ((image.anchor._from.row, image.anchor._from.col, image) for image in sheet._images),
                                key=lambda x: (x[0], x[1])
                                )
                                                    
                                
                                
                                #Track col number (see just below)
                                col_number=-1

                                # Initialize database (where images metadata will be stored)
                                image_database = []
                                # Dictionary to store PIL images in memory with unique IDs
                                image_memory = {}

                                for idx, (row, col, image) in enumerate(images_with_positions):
                                    
                                    #Track col number, as some photos are in the same col so we have to increment
                                    #it for the uniqueness of image names
                                    if col_number==col:
                                        col_number+=1
                                    else:
                                        col_number=col
                                    # Extract the binary content of the image
                                    img_bytes = image.ref.getvalue()  # Get the image binary data
                                    
                                    # Convert to a PIL image for display
                                    pil_image = PILImage.open(BytesIO(img_bytes))
                                    
                                    
                                    

                                    # Retrieve textual information 
                                    if row > 1:  # Ensure the row exists (skip if it's the first row)
                                        
                                        PB_data = [
                                            sheet.cell(row=row-3 , column=c).value
                                            for c in range(1, sheet.max_column + 1)
                                        ]
                                        Address_data = [
                                            sheet.cell(row=row , column=c).value
                                            for c in range(1, sheet.max_column + 1)
                                        ]

                                        
                                        #st.write(f"Textual Information (Row {row - 3}):", PB_data)
                                        #st.write(f"Textual Information (Row {row }):", Address_data)
                                    
                                    
                                    # Save the image to the directory with a unique name based on row and column
                                    
                                    #Replace spaces(not convenient) and slashes(not allowed) with "_" as they are not allowed in file names
                                    id_address=Address_data[3].replace(" ","_").replace("/","_").replace("\t","")
                                    
                                    img_title=f"image_{row}_{col_number}_{id_address}.png"
                                    img_filename = os.path.join(save_directory, img_title)
                                    
                                    
                                    # Store image in memory with its unique ID
                                    image_memory[img_title] = pil_image

                                    

                                    if image_save:
                                        pil_image.save(img_filename)  # Save as PNG


                                    ##Save Image info to DB

                                    # Flatten PB_Data and Address_data into a single dictionary
                                    flattened_data = {}
                                    
                                    #Get Planche number from the file title
                                    planche=get_planche_number(uploaded_files.name)

                                    # Add the image identifier and metadata into the list
                                    flattened_data["id"] = img_title  # Add the image name as an identifier
                                    flattened_data["planche"]= planche
                                    
                                    # Flatten PB_Data and Address_data into a single dictionary
                                    flattened_data.update({
                                        key.replace(":", "").strip(): value.strip()
                                        for key, value in zip(PB_data[::2], PB_data[1::2])
                                    })
                                    flattened_data.update({
                                        key.replace(":", "").strip(): value.strip()
                                        for key, value in zip(Address_data[::2], Address_data[1::2])
                                    })

                                    image_database.append(flattened_data)

                                                                    
                                    # Display Image data
                                    #st.json(flattened_data)
                                
                                
                                #Transform the db into a structure avoiding redundancy
                                # Group entries by their unique metadata values
                                grouped_db = {}
                                for entry in image_database:
                                    # Create a unique key based on metadata values
                                    metadata_key = (entry["planche"], entry["PB"], entry["Emplacement"], entry["Adresse"], entry["Site Support"])
                                    
                                    # If the group doesn't exist, create it
                                    if metadata_key not in grouped_db:
                                        grouped_db[metadata_key] = []
                                    
                                    # Append the image ID to the group
                                    grouped_db[metadata_key].append({"id": entry["id"]})

                                # Convert the grouped database into the new structured format
                                new_db = []
                                for metadata_key, images in grouped_db.items():
                                    metadata = {
                                        "planche": metadata_key[0],
                                        "PB": metadata_key[1],
                                        "Emplacement": metadata_key[2],
                                        "Adresse": metadata_key[3],
                                        "Site Support": metadata_key[4]
                                    }
                                    new_db.append({
                                        "metadata": metadata,
                                        "images": images
                                    })


                                # Save the database as a JSON file
                                database_path = os.path.join("data", "image_database.json")
                                with open(database_path, "w", encoding="utf-8") as json_file:
                                    json.dump(new_db, json_file, ensure_ascii=False, indent=4)

                                if image_save:
                                    st.toast(f"Les photos ont été récupérées et sauvegardés sur: {save_directory}")
                                else:
                                    st.toast(f"Les photos ont été récupérées avec succès!" ,icon="✅")

                                    # Display the image
                                    #st.write(f"Image at Row: {row}, Column: {col}")
                                    #st.image(pil_image)
                                
                                #foa_feeder(new_db,"00_C16.xlsx",image_memory,outputs_directory)#desktop
                                foa_feeder(new_db,"00_C16.xlsx",command_num,image_memory,base_filename)#web
                                
                                if "zip_ready" in st.session_state and st.session_state["zip_ready"]:
                                    zip_file_path = st.session_state["zip_file_path"]

                                    if os.path.exists(zip_file_path):
                                        with open(zip_file_path, "rb") as zipf:
                                            zip_data = zipf.read()  # Read the file before passing it to the button
                                            
                                            st.download_button(
                                                label="📥 Téléchargez vos fichiers",
                                                data=zip_data,  # Pass the file content, not the open file
                                                file_name=os.path.basename(zip_file_path),
                                                mime="application/zip",
                                                key="download_zip"  # Unique key to prevent disappearance
                                            )

                                        # Ensure the button remains visible after first display
                                        st.session_state["show_download"] = True

                                    # if st.session_state.get("show_download", False):
                                    #     st.success("✅ Vos fichiers sont prêts! Cliquez ci-dessus pour télécharger.")
                            except Exception as e:
                                st.error(f"An error occurred while processing the file: {e}")
                    
            #else:
                #st.warning("Merci de charger un fichier Excel")
    
    
    elif app == "Annexe C6":


        
         # Create Tabs for Extraction and Insertion
        extra_tab,inser_tab= st.tabs(["🔍 Extraction", "📝 Insertion"])

        with extra_tab:
            st.subheader("🔍 Extraction")
        

            #st.markdown(f'<h1 style="color:#ffa500;font-size:44px;">{"Annexe C6"}</h1>', unsafe_allow_html=True)

            # Inject the custom CSS to override Streamlit file uploader
            st.markdown(upload_style,unsafe_allow_html=True,)
            st.write(css, unsafe_allow_html=True)


            #Uploading the file
            uploaded_filesC6=upload_files(type="xlsx",multiple=False)

            

            if uploaded_filesC6:
                #Getting workbook from the session, otherwise st keeps refreshing and reload 
                #the wb each time it refreshes, and decreases user exp quality
                if "workbookC6" not in st.session_state:
                    with st.spinner("Lecture et validation du fichier..."):
                        wbC6 = load_excel_file(uploaded_filesC6,read_only=True)  # Load the Excel file
                        st.session_state.workbookC6 = wbC6  # Store the workbook in session state
                        st.session_state.sheet_names = wbC6.sheetnames  # Store sheet names

                # Access stored workbook and sheet names
                wbC6 = st.session_state.workbookC6
                sheet_names = st.session_state.sheet_names


                base_filename = os.path.splitext(uploaded_filesC6.name)[0]

                # Add an empty default choice (So the user is encouraged to correctly choose
                #the correct sheet name, otherwise maybe he will inadvertly let the first option
                #on the menu, which is not surely the good one)
                sheet_options = [""] + sheet_names  # Prepend an empty choice

                #Get nmero commande
                #command_num=st.text_input("Merci de renseigner le numéro de commande")        
                
                
                # Step 2: Prompt for the sheet name and provide a button to process
                sheet_name = st.selectbox("Veuillez sélectionner la feuille contenant les photos puis cliquez sur Traiter", sheet_options)
                process_button = st.button("Traiter")

                wbC6.close()  # Close the workbook to free up memory

                #What happens when "Traiter" is clicked
                if process_button:
                    # Check if the user has selected a valid sheet name
                    if sheet_name == "":
                        st.warning("Veuillez sélectionner une feuille avant de continuer.")
                    else:
                        with st.spinner("Traitement en cours..."):
                            try:
                                
                                
                                zip_filename = extract_images_with_correct_names(uploaded_filesC6, base_filename)

                                # Step 4: Provide a download link for the ZIP file
                                with open(zip_filename, "rb") as f:
                                    st.download_button(
                                        label="Téléchargez vos photos!",
                                        data=f,
                                        file_name="extracted_images.zip",
                                        mime="application/zip"
                                    )




                            except Exception as e:
                                    st.error(f"An error occurred while processing the file: {e}")


        with inser_tab:
            st.subheader("📝 Insertion")
            
            # Step 2: User uploads the Excel file
            uploaded_file = st.file_uploader("Téléchargez le fichier Excel", type=["xlsx"])

            # Step 2: Upload multiple images
            uploaded_photos = st.file_uploader("📸 Chargez les photos associées", type=["png", "jpg", "jpeg"], accept_multiple_files=True)
            if uploaded_file and uploaded_photos:
                # Step 2: Get the directory where the uploaded file is stored
                # file_name = uploaded_file.name  # Get the uploaded file's name
                # abs_path=os.path.abspath(file_name)
                # base_folder =  os.path.dirname(abs_path) # get folder path
                # folder_path = os.path.join("Photos", base_folder)  # append Photos
                # Step 1: Save the uploaded file to a temporary location
                # Step 1: Get the name of the uploaded file (only the filename, no path)
                

                with st.spinner("Analyse du fichier Excel..."):
                    # Step 3: Extract photo names from Excel
                    photo_names = extract_photo_names_from_excel(uploaded_file)

                    # Step 4: Filter photo names based on rules
                    filtered_names = filter_photo_names(photo_names)

                    # Step 5: Find matching photos in the folder
                    selected_photos = find_matching_photos(uploaded_photos, filtered_names)

                # Step 6: Display results
                st.success(f"{len(selected_photos)} photos sélectionnées!")

                with open("temp.xlsx", "wb") as f:
                    f.write(uploaded_file.getbuffer())

                with st.spinner("Traitement en cours..."):
                # Process the Excel file
                    process_excel("temp.xlsx",selected_photos)

                # Provide a download link for the updated file
                with open("temp.xlsx", "rb") as f:
                    st.download_button(
                        label="Download Updated Excel File",
                        data=f,
                        file_name="updated_file.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )

                # Show images
                # for photo in selected_photos:
                #     try:
                #         img = PILImage.open(photo)
                #         st.image(img, caption=photo.name, use_column_width=True)
                #     except Exception as e:
                #         st.error(f"❌ Impossible d'afficher {photo.name}: {e}")


    
            
        




if __name__ == "__main__":
    main()



