
import openpyxl as px
import os
import re
import streamlit as st
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
import io
from PIL import Image as PILImage


def extract_photo_names_from_excel(excel_file):
    """Extracts photo names ending with _1, _2, _3, or _4 from the Excel file."""
    wb = px.load_workbook(excel_file, read_only=True)
    sheet = wb["Photos"]  

    photo_names = set()

    for row in sheet.iter_rows(values_only=True):
        for cell in row:
            if isinstance(cell, str) and cell.endswith(("_1.jpg", "_2.jpg", "_3.jpg", "_4.jpg")):
                photo_names.add(cell.strip())  # Ensure there are no spaces

    return photo_names



def filter_photo_names(photo_names):
    """Filters the photo names according to the rule: if _3 and _4 exist, ignore all 4."""
    grouped_names = {}

    for name in photo_names:
        base_name = name[:-6]  # Remove the last 2 characters (_1, _2, _3, _4)
        if base_name not in grouped_names:
            grouped_names[base_name] = set()
        grouped_names[base_name].add(name[-6:])  # Store _1, _2, _3, _4 suffixes

    filtered_names = set()

    for base_name, suffixes in grouped_names.items():
        if "_3.jpg" in suffixes and "_4.jpg" in suffixes:
            continue  # Ignore all if _3 and _4 exist
        filtered_names.update({base_name + suffix for suffix in suffixes})  # Keep valid ones

    return filtered_names



def find_matching_photos(uploaded_photos, filtered_names):
    """Finds photos in the given folder that match filtered names and end with _3 or _4."""
    selected_photos = []

    # Extract base names from filtered_names (removing _1, _2, _3, _4 and file extension)
    base_names = {re.sub(r"_\d+\.jpg$", "", name) for name in filtered_names}

    
    
    for filename in uploaded_photos:
        
        name_without_ext = os.path.splitext(filename.name)[0]  # Remove file extension
        
        # Ensure filename has the correct format (must end with _3 or _4)
        base_name_match = re.match(r"(.+)_([34])$", name_without_ext)

        if base_name_match:
            base_name, suffix = base_name_match.groups()

            # Condition: Base name must be in base_names AND suffix must be _3 or _4
            if base_name in base_names:
                selected_photos.append(filename)
    
    return selected_photos

# Function to get the size of the photo in cell A2
def get_photo_size_in_A2(sheet):
    # Access cell A2
    cell_A2 = sheet['A2']
    if cell_A2.value and isinstance(cell_A2.value, Image):
        # Get the dimensions of the photo in A2
        return cell_A2.value.width, cell_A2.value.height
    return None  # No photo found in A2

# Function to resize all selected photos to a fixed size
def resize_photos_to_fixed_size(selected_photos, target_width_cm, target_height_cm):
    # Convert centimeters to pixels (1 cm ≈ 37.795 pixels)
    target_width_px = int(target_width_cm * 37.795)
    target_height_px = int(target_height_cm * 37.795)

    resized_photos = []
    for photo in selected_photos:
        # Open the image using PIL
        pil_img = PILImage.open(photo)
        # Resize the image to the target size
        resized_img = pil_img.resize((target_width_px, target_height_px), PILImage.ANTIALIAS)
        # Convert the resized image back to bytes
        img_bytes = io.BytesIO()
        resized_img.save(img_bytes, format="PNG")
        img_bytes.seek(0)
        resized_photos.append(img_bytes)
    return resized_photos



# Function to process the Excel file
def process_excel(file_path,selected_photos):
    # Load the workbook and select the active sheet
    workbook = px.load_workbook(file_path)
    sheet = workbook["Photos"]

    # Iterate through each row in the sheet
    for row in sheet.iter_rows(min_row=2, max_col=2, values_only=False):  # Skip header row
        # Get the values from the first and second columns
        col1_value = row[0].value
        col2_value = row[1].value

        # Check if the filenames end with '_1.jpg' and '_2.jpg'
        if col1_value and col2_value and col1_value.endswith('_1.jpg') and col2_value.endswith('_2.jpg'):
            # Extract the base name (without '_1.jpg' or '_2.jpg')
            base_name = col1_value[:-6]  # Remove '_1.jpg' from the end

            # Generate new filenames with '_3.jpg' and '_4.jpg'
            new_filename_3 = f"{base_name}_3.jpg"
            new_filename_4 = f"{base_name}_4.jpg"

            # Write the new filenames to the third and fourth columns
            sheet.cell(row=row[0].row, column=3, value=new_filename_3)
            sheet.cell(row=row[0].row, column=4, value=new_filename_4)
   
   
   
    # Define the fixed size in centimeters
    target_width_cm = 8.47  # Width in cm
    target_height_cm = 10.58  # Height in cm

    # Resize all selected photos to the fixed size
    resized_photos = resize_photos_to_fixed_size(selected_photos, target_width_cm, target_height_cm)

    # Iterate through each selected photo
    for photo, resized_photo in zip(selected_photos, resized_photos):
        # Get the photo name (without path)
        photo_name = os.path.basename(photo.name)

        # Search for the cell containing the photo name
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value == photo_name:
                    # Found the cell with the photo name
                    # Insert the resized image into the cell above
                    img = Image(resized_photo)
                    sheet.add_image(img, anchor=cell.offset(row=-1, column=0).coordinate)  # Insert above the cell
                    break



    # Save the updated workbook
    workbook.save(file_path)
    st.success("Excel file processed and saved successfully!")


