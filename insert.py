import datetime
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.drawing.xdr import XDRPoint2D, XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU, cm_to_EMU
import os
#import win32com.client as win32
from openpyxl.utils import range_boundaries
import xlsxwriter
import openpyxl as px
import streamlit as st
import zipfile

# # Paths to the files
# image_path = "D:\Kamel\IA_Projects\FOA_Builder\saved_images\AXTD_FI-48092-0005_PA-48092-000C_MADSITE/image_214_0_CH_FT_341_L1T.png"  # Path to the saved image
# file_path = "00_C16.xlsx"  # Path to the Excel file where the image will be inserted
# save_path= "00_C16_updated.xlsx"  # Path to save the updated Excel file

# # Predefined location
# target_sheet_name = "CH"  # Name of the target sheet
# target_cell = "A6"  # Cell where the image will be placed

# # Load the output Excel file
# with open(file_path, "rb") as f:
#     wb = load_workbook(f)

# if target_sheet_name not in wb.sheetnames:
#     raise ValueError(f"Sheet '{target_sheet_name}' not found in the workbook.")

# sheet = wb[target_sheet_name]



# # Insert the image
# if os.path.exists(image_path):
#     img = Image(image_path)
#     img.width, img.height = 365, 220  # Optionally resize the image (width, height in pixels)
#     sheet.add_image(img, target_cell)
#     print(f"Image inserted at {target_cell} on '{target_sheet_name}'")
# else:
#     print(f"Image not found: {image_path}")

# # Save the modified Excel file
# wb.save(save_path)
# wb.close()
# print(f"Updated Excel file saved at: {save_path}")



# ##########FOR XLSM##########

# # Define file paths
# image_path = "D:\Kamel\IA_Projects\FOA_Builder\saved_images\AXTD_FI-48092-0005_PA-48092-000C_MADSITE/image_214_0_CH_FT_341_L1T.png"  # Path to the saved image
# file_path = "00_C16.xlsm"  # Path to the Excel file where the image will be inserted
# save_path= "00_C16_updated.xlsm"  # Path to save the updated Excel file



# # Open Excel
# excel = win32.gencache.EnsureDispatch("Excel.Application")
# excel.Visible = False  # Keep Excel hidden during the process

# # Open the workbook
# workbook = excel.Workbooks.Open(file_path)

# # Select the sheet where the image will be inserted
# sheet = workbook.Sheets("Sheet1")  # Change "Sheet1" to your desired sheet name

# # Insert the image (adjust placement as needed)
# sheet.Shapes.AddPicture(
#     image_path,  # Path to the image
#     LinkToFile=False,
#     SaveWithDocument=True,
#     Left=100,  # Horizontal position in points
#     Top=100,   # Vertical position in points
#     Width=200, # Image width in points
#     Height=150 # Image height in points
# )

# # Save the file
# workbook.SaveAs(save_path, FileFormat=52)  # FileFormat=52 is for .xlsm files
# workbook.Close()
# excel.Quit()


# def foa_feeder(db, template_file, pil_images):
    
#     target_sheet_name = "CH"  # Name of the target sheet

#      # Predefined locations and sizes for each image
#     image_positions = [
#         {"cell": "A6", "width": 365, "height": 220},  # First image
#         {"cell": "N35", "width": 390, "height": 225},  # Second image
#         {"cell": "C74", "width": 234, "height": 186},  # Third image
#         {"cell": "L74", "width": 234, "height": 186},  # Fourth image
#         {"ell": "U74", "width": 234, "height": 186},  # Third image
#         {"cell": "AD74", "width": 234, "height": 186},  # Fourth image
#     ]

    
#    # Load the template Excel file once
#     with open(template_file, "rb") as f:
#         template_data = f.read()

#     for entry in db:
#         try:
#             # Reload the workbook from the template for each entry
#             wb = load_workbook(filename=template_file)
#             if target_sheet_name not in wb.sheetnames:
#                 raise ValueError(f"Sheet '{target_sheet_name}' not found in the workbook.")

#             sheet = wb[target_sheet_name]

#             # Extract metadata
#             img_planche = entry['metadata']['planche']
#             img_address = entry['metadata']['Adresse']
#             img_chambre = entry['metadata']['Site Support'].split(" ")[1]
#             img_type = entry['metadata']['Site Support'].split(" ")[2]
#             img_supp = entry['metadata']['Site Support']

#             images = entry['images']

#             # Extract the list of image ids
#             image_ids = [image["id"] for image in images]
#             # print(f"Processing image IDs: {image_ids}")

#             # Get the actual PIL images
#             current_pil_images = [pil_images[id] for id in image_ids]

#             # Unique output file and temp image path
#             output_file = f"{img_chambre}_C16_updated.xlsx"


#             # Insert each image based on predefined positions and sizes
#             for idx, pil_image in enumerate(current_pil_images):
#                 if idx >= len(image_positions):
#                     print(f"No predefined position for image {idx + 1}, skipping.")
#                     continue

#                 # Get the location and size for the current image
#                 image_position = image_positions[idx]
#                 cell = image_position["cell"]
#                 width = image_position["width"]
#                 height = image_position["height"]

#                 # Save the PIL image temporarily
#                 temp_image_path = f"temp_image_{img_chambre}_{idx}.png"
#                 try:
#                     pil_image.save(temp_image_path)
#                 except Exception as e:
#                     print(f"Failed to save image {idx + 1} for chambre {img_chambre}: {e}")
#                     continue

#                 # Insert the image into the Excel sheet
#                 if os.path.exists(temp_image_path):
#                     img = Image(temp_image_path)
#                     img.width, img.height = width, height  # Set the image size
                    
                    
#                     sheet.add_image(img, cell)
#                     print(f"Image inserted at {cell} on '{target_sheet_name}' with size ({width}x{height})")
#                     #os.remove(temp_image_path)  # Remove temporary image after use
#                 else:
#                     print(f"Temp image not found: {temp_image_path}")

#             # Save the modified Excel file
#             wb.save(output_file)
#             print(f"Updated Excel file saved at: {output_file}")

            
#             # Close the workbook to release resources
#             wb.close()

#         except Exception as e:
#             print(f"Error processing entry: {e} for chambre: {img_chambre}")

#     st.info("Vos fichiers sont prêts.")





def foa_feeder(db, template_file, pil_images,output_dr):
    
    target_sheet_name = "CH"  # Name of the target sheet

    # Predefined locations and sizes for each image
    image_positions = [
        {"cell": "A6", "width": 365, "height": 220},  # First image
        {"cell": "M33", "width": 380, "height": 225},  # Second image
        {"cell": "C74", "width": 234, "height": 186},  # Third image
        {"cell": "L74", "width": 234, "height": 186},  # Fourth image
        {"cell": "U74", "width": 234, "height": 186},  # Fifth image
        {"cell": "AD74", "width": 234, "height": 186},  # Sixth image
    ]
    
    

    # Conversion functions (cm to EMUs)
    c2e = lambda x: int(x * 360000)  # Conversion factor for cm to EMU (English Metric Units)
    cellh = lambda x: c2e((x * 49.77) / 99)  # Row height in EMUs
    cellw = lambda x: c2e((x * (18.65 - 1.71)) / 10)  # Column width in EMUs

    # List to store paths of temporary images
    temp_image_paths = []
    for entry in db:
        try:
            # Reload the workbook from the template for each entry
            wb = load_workbook(filename=template_file)
            if target_sheet_name not in wb.sheetnames:
                raise ValueError(f"Sheet '{target_sheet_name}' not found in the workbook.")

            sheet = wb[target_sheet_name]

            # Extract metadata
            img_planche = entry['metadata']['planche']
            img_address = entry['metadata']['Adresse']
            img_chambre = entry['metadata']['Site Support'].split(" ")[1]
            img_type = entry['metadata']['Site Support'].split(" ")[2]
            img_supp = entry['metadata']['Site Support']

            images = entry['images']

            # Insert values into specific cells
            sheet["T2"] = img_chambre      
            sheet["Y2"] = img_planche          
            sheet["AD2"] = img_type    
            sheet["AI2"] = datetime.datetime.today().strftime('%d/%m/%Y') 
            sheet["T3"] = img_address          



            # Extract the list of image ids
            image_ids = [image["id"] for image in images]

            # Get the actual PIL images
            current_pil_images = [pil_images[id] for id in image_ids]

             #Add this if Web app
            outputs_directory = os.path.join("Outputs", output_dr)

            # Unique output file and temp image path
            
            #output_file = f"{output_dr}\{img_chambre}_C16_updated.xlsx"
            #output_file = os.path.join(outputs_directory, f"{img_chambre}_C16.xlsx") #Desktop
            output_file = os.path.join(outputs_directory, f"{img_chambre}_C16.xlsx") #to work in any OS

            # Insert each image based on predefined positions and sizes
            for idx, pil_image in enumerate(current_pil_images):
                if idx >= len(image_positions):
                    print(f"No predefined position for image {idx + 1}, skipping.")
                    continue

                # Get the location and size for the current image
                image_position = image_positions[idx]
                cell = image_position["cell"]
                width = image_position["width"]
                height = image_position["height"]

                # Save the PIL image temporarily
                temp_image_path = f"temp_image_{img_chambre}_{idx}.png"
                try:
                    pil_image.save(temp_image_path)
                except Exception as e:
                    print(f"Failed to save image {idx + 1} for chambre {img_chambre}: {e}")
                    continue
                
                temp_image_paths.append(temp_image_path)

                # Insert the image into the Excel sheet
                if os.path.exists(temp_image_path):
                    img = Image(temp_image_path)
                    img.width, img.height = width, height  # Set the image size

                    # Move the image if it is the second image (idx == 1)
                    if idx == 1:  # Move only the second image
                        # Slightly move to the right (0.5 columns) and down (0.5 rows)
                        column = 12  # Column 'C'
                        row = 32  # Row 6 in Excel (index 5 in 0-based index)
                        coloffset = cellw(0.3)  # Offset 0.5 column to the right
                        rowoffset = cellh(0.7)  # Offset 0.5 row down


                        # Image size (in pixels) converted to EMUs (1 pixel = 9525 EMUs)
                        
                        image_width = img.width * 9525
                        image_height = img.height * 9525

                        # Create the AnchorMarker with offset values
                        marker = AnchorMarker(col=column, colOff=coloffset, row=row, rowOff=rowoffset)

                        # Create the XDRPositiveSize2D for the size of the image
                        ext = XDRPositiveSize2D(cx=image_width, cy=image_height)

                        # Apply the anchor and size to the image
                        img.anchor = OneCellAnchor(_from=marker, ext=ext)

                        sheet.add_image(img)
                    else:
                        # Add the image to the sheet
                        sheet.add_image(img, cell)

                    
                    print(f"Image inserted at {cell} on '{target_sheet_name}' with size ({width}x{height})")
                else:
                    print(f"Temp image not found: {temp_image_path}")

            # Save the modified Excel file
            wb.save(output_file)
            print(f"Updated Excel file saved at: {output_file}")

            # Close the workbook to release resources
            wb.close()

        except Exception as e:
            print(f"Error processing entry: {e} for chambre: {img_chambre}")

    for temp_image_path in temp_image_paths:
        try:
            os.remove(temp_image_path)
            #print(f"Deleted: {temp_image_path}")
        except FileNotFoundError:
            print(f"File already deleted or not found: {temp_image_path}")
        except Exception as e:
            print(f"Error deleting file {temp_image_path}: {e}")


   
    # Create a ZIP file containing all output Excel files
    zip_filename = f"{output_dr}.zip"

    with zipfile.ZipFile(zip_filename, "w") as zipf:
        for file in os.listdir(outputs_directory):
            file_path = os.path.join(outputs_directory, file)
            if file != f"{output_dr}.zip":
                zipf.write(file_path, os.path.basename(file_path))  # Add only valid files

    # Provide a download button for the ZIP file
    with open(zip_filename, "rb") as zipf:
        st.download_button(
            label="📥 Téléchargez vos fichiers",
            data=zipf,
            file_name=f"{output_dr}.zip",
            mime="application/zip"
        )

    st.info("Vos fichiers sont prêts!")



